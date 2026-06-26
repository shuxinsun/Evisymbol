#!/usr/bin/env python3
import os
import re
from collections import defaultdict
from openpyxl import Workbook

LOG_ROOT = "../weight-logs"
OUTPUT_FILE = "./weight.txt"
OUTPUT_XLSX = "./weight.xlsx"

addr_keys = [
    "in_range", "out_0_1kb", "out_1_2kb",
    "out_2_3kb", "out_3_4kb", "out_4kb_plus"
]

align_keys = [
    "total", "aligned_8", "aligned_4",
    "aligned_2", "misaligned"
]

addr_stats = defaultdict(lambda: defaultdict(int))
align_stats = defaultdict(lambda: defaultdict(int))
evi_imm = defaultdict(lambda: defaultdict(int))  # Pred
evi_sym = defaultdict(lambda: defaultdict(int))  # GT

total_addr = defaultdict(int)
total_align = defaultdict(int)
total_evi_imm = defaultdict(int)
total_evi_sym = defaultdict(int)


def extract_section(lines, header):
    for i, line in enumerate(lines):
        if header in line:
            return lines[i:]
    return []


folders = [
    f for f in os.listdir(LOG_ROOT)
    if os.path.isdir(os.path.join(LOG_ROOT, f))
    and os.path.isfile(os.path.join(LOG_ROOT, f, "log.txt"))
]

total_logs = len(folders)

print("Found {} log folders.\n".format(total_logs))

for idx, folder in enumerate(sorted(folders), start=1):
    path = os.path.join(LOG_ROOT, folder)
    log_file = os.path.join(path, "log.txt")

    opt = folder.split("-")[1]

    percent = (idx / total_logs) * 100
    print("[{}/{}] ({:.1f}%) Processing: {}".format(idx, total_logs, percent, folder))

    with open(log_file, "r", encoding="utf-8", errors="ignore") as f:
        lines = f.readlines()

    section = extract_section(lines, "GT Address Range Stats:")
    for key in addr_keys:
        for line in section:
            if line.strip().startswith(key):
                val = int(line.split()[1])
                addr_stats[opt][key] += val
                total_addr[key] += val
                break

    section = extract_section(lines, "GT Data Alignment Stats:")
    for key in align_keys:
        for line in section:
            if line.strip().startswith(key):
                val = int(line.split()[1])
                align_stats[opt][key] += val
                total_align[key] += val
                break

    in_evi = False
    for line in lines:
        if "Evidence statistics" in line:
            in_evi = True
            continue
        if in_evi:
            if line.strip() == "" or line.startswith("-"):
                continue
            if line.startswith("GT"):
                break
            match = re.match(r"(r\d+[-\w]*)\s+(\d+)\s+(\d+)", line)
            if match:
                name = match.group(1)
                imm = int(match.group(2))  # Pred
                sym = int(match.group(3))  # GT

                evi_imm[opt][name] += imm
                evi_sym[opt][name] += sym

                total_evi_imm[name] += imm
                total_evi_sym[name] += sym

# ========================
# 原 txt 输出（保持不变）
# ========================
with open(OUTPUT_FILE, "w") as out:

    out.write("===== Address Range =====\n")
    for opt in sorted(addr_stats.keys()):
        out.write(f"\n[{opt}]\n")
        for key in addr_keys:
            out.write(f"{key}: {addr_stats[opt][key]}\n")

    out.write("\n[TOTAL]\n")
    for key in addr_keys:
        out.write(f"{key}: {total_addr[key]}\n")

    out.write("\n\n===== Alignment =====\n")
    for opt in sorted(align_stats.keys()):
        out.write(f"\n[{opt}]\n")
        for key in align_keys:
            out.write(f"{key}: {align_stats[opt][key]}\n")

    out.write("\n[TOTAL]\n")
    for key in align_keys:
        out.write(f"{key}: {total_align[key]}\n")

    out.write("\n\n===== Evidence =====\n")
    for opt in sorted(evi_imm.keys()):
        out.write(f"\n[{opt}]\n")
        for name in sorted(evi_imm[opt].keys()):
            imm = evi_imm[opt][name]
            sym = evi_sym[opt][name]
            ratio = (sym / imm * 100) if imm > 0 else 0
            out.write(f"{name}: imm={imm}, sym={sym}, ratio={ratio:.2f}%\n")

print("\nTXT statistics written to:", OUTPUT_FILE)


wb = Workbook()
ws = wb.active
ws.title = "Summary"

headers = ["优化级别", "统计类型", "AR", "ST", "RE", "CE", "JE", "FE", "PR", "PLT", "AE"]
ws.append(headers)

mapping = {
    "AR": "r1-address-range",
    "ST": "r2-symtab",
    "RE": "r3-relocation-exact",
    "CE": "r4-call",
    "JE": "r5-jmp-reference",
    "FE": "r6-function-entry",
    "PR": "r13-pc-relative",
    "PLT": "r12-plt-entry",
    "AE": "r17-aligned"
}

opt_levels = ["O0", "O1", "O2", "O3", "Os", "Ofast"]

def write_ratio_row(label1, label2, pred_dict, gt_dict):
    row = [label1, label2]
    ws.append(row)
    row_index = ws.max_row

    col_index = 3
    for key in mapping:
        pred = pred_dict.get(mapping[key], 0)
        gt = gt_dict.get(mapping[key], 0)
        ratio = (gt / pred * 100) if pred > 0 else 0
        cell = ws.cell(row=row_index, column=col_index, value=ratio)
        cell.number_format = "0.00"
        col_index += 1


for opt in opt_levels:

    row_pred = [opt, "Pred"]
    for key in mapping:
        row_pred.append(evi_imm[opt].get(mapping[key], 0))
    ws.append(row_pred)

    row_gt = ["", "GT"]
    for key in mapping:
        row_gt.append(evi_sym[opt].get(mapping[key], 0))
    ws.append(row_gt)

    write_ratio_row("", "%", evi_imm[opt], evi_sym[opt])


row_total_pred = ["TOTAL", "Pred"]
for key in mapping:
    row_total_pred.append(total_evi_imm.get(mapping[key], 0))
ws.append(row_total_pred)

row_total_gt = ["", "GT"]
for key in mapping:
    row_total_gt.append(total_evi_sym.get(mapping[key], 0))
ws.append(row_total_gt)

write_ratio_row("", "%", total_evi_imm, total_evi_sym)


wb.save(OUTPUT_XLSX)

print("Excel statistics written to:", OUTPUT_XLSX)
print("\nAll logs processed.")
